#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сырая выгрузка сделок для просмотра (проект «проверка тегов и источника», этап 1).

Ничего не меняет в amoCRM. Достаёт сделки за окно и кладёт их плоской таблицей в
.xlsx — чтобы глазами посмотреть реальные данные: какие бывают источники, теги,
как заполнены даты, и решить, как их анализировать.

Колонки заточены под задачу: ID, ссылка, воронка/этап, ответственный,
дата создания, дата «вступил в чат», источник, utm-метки, теги.

Запуск локально (нужен только AMO_TOKEN в .env):
    python dump_source_tags.py                 # по умолчанию последние 14 дней
    python dump_source_tags.py --days 30
    python dump_source_tags.py --months 3      # как окно основной выгрузки

Клиент amoCRM и разбор полей переиспользуются из amo_export.py.
"""

import os
import sys
import argparse
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo


# --- подхватываем .env до импорта amo_export (он читает AMO_TOKEN на импорте) ---
def _load_env_file():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    _load_env_file()

from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import amo_export as ax


OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "analysis")

COLUMNS = [
    "ID", "Ссылка", "Название", "Воронка", "Этап", "Ответственный",
    "Дата создания", "Дата вступил в чат",
    "Источник", "utm_source", "utm_medium", "utm_campaign", "Теги",
]


def parse_args():
    p = argparse.ArgumentParser(description="Сырая выгрузка сделок amoCRM для просмотра")
    g = p.add_mutually_exclusive_group()
    g.add_argument("--months", type=int, help="окно в месяцах до вчера")
    g.add_argument("--days", type=int, default=14, help="окно в днях до вчера (по умолчанию 14)")
    return p.parse_args()


def ts_to_ddmmyyyy(ts):
    """Unix-секунды → 'dd.mm.yyyy' по Москве."""
    try:
        num = float(ts)
    except (TypeError, ValueError):
        return ""
    if not num or num < 1000000000:
        return ""
    d = datetime.fromtimestamp(num, tz=timezone.utc) + timedelta(hours=3)
    return d.strftime("%d.%m.%Y")


def serial_to_ddmmyyyy(serial):
    """Серийная дата Google Sheets (целое) → 'dd.mm.yyyy'. Пустое — как есть."""
    if serial in (None, ""):
        return ""
    try:
        s = int(float(serial))
    except (TypeError, ValueError):
        return str(serial)
    try:
        d = datetime(1899, 12, 30) + timedelta(days=s)
        return d.strftime("%d.%m.%Y")
    except Exception:
        return str(serial)


def main():
    args = parse_args()

    if not ax.AMO_TOKEN:
        print("ОШИБКА: не задан AMO_TOKEN. Заполни .env (см. .env.example).")
        sys.exit(1)

    msk = ZoneInfo(ax.TIMEZONE)
    yesterday = datetime.now(msk) - timedelta(days=1)
    date_to_dt = yesterday.replace(hour=23, minute=59, second=59, microsecond=0)
    if args.months:
        date_from_dt = (yesterday - relativedelta(months=args.months)).replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        date_from_dt = (yesterday - timedelta(days=args.days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)
    date_from_ts = int(date_from_dt.timestamp())
    date_to_ts = int(date_to_dt.timestamp())

    print(f"Окно: {date_from_dt:%d.%m.%Y} — {date_to_dt:%d.%m.%Y} (created_at, МСК)")
    print(f"Воронки: {ax.PIPELINE_IDS}")

    # справочники: пользователи (имена) и воронки/этапы
    users = ax.amo_fetch_all("/api/v4/users", {"limit": 250}, "users")
    user_map = {str(u["id"]): (u.get("name") or "") for u in users}
    pipelines_data = ax.amo_get("/api/v4/leads/pipelines")
    pipelines = (pipelines_data.get("_embedded") or {}).get("pipelines") or []
    pipeline_map, status_map = {}, {}
    for p in pipelines:
        pid = str(p["id"])
        pipeline_map[pid] = p.get("name") or ""
        for s in ((p.get("_embedded") or {}).get("statuses") or []):
            status_map[str(s["id"])] = s.get("name") or ""

    # сделки (без контактов)
    print("Тяну сделки...")
    params = {
        "limit": ax.AMO_PAGE_LIMIT,
        "filter[created_at][from]": date_from_ts,
        "filter[created_at][to]": date_to_ts,
        "order[created_at]": "asc",
    }
    for i, pid in enumerate(ax.PIPELINE_IDS):
        params[f"filter[pipeline_id][{i}]"] = pid
    leads = ax.amo_fetch_all("/api/v4/leads", params, "leads")

    allowed = set(ax.PIPELINE_IDS)
    leads = [l for l in leads
             if date_from_ts <= int(l.get("created_at") or 0) <= date_to_ts
             and int(l.get("pipeline_id") or 0) in allowed]
    print(f"  сделок: {len(leads)}")

    # строки
    rows = []
    for l in leads:
        lid = l.get("id") or ""
        rows.append({
            "ID": lid,
            "Ссылка": f"{ax.AMO_BASE_URL}/leads/detail/{lid}" if lid else "",
            "Название": l.get("name") or "",
            "Воронка": pipeline_map.get(str(l.get("pipeline_id")), l.get("pipeline_id") or ""),
            "Этап": status_map.get(str(l.get("status_id")), l.get("status_id") or ""),
            "Ответственный": user_map.get(str(l.get("responsible_user_id")), l.get("responsible_user_id") or ""),
            "Дата создания": ts_to_ddmmyyyy(l.get("created_at")),
            "Дата вступил в чат": serial_to_ddmmyyyy(ax.get_lead_field(l, "Дата вступил в чат")),
            "Источник": ax.get_lead_field(l, "Источник"),
            "utm_source": ax.get_lead_field(l, "utm_source"),
            "utm_medium": ax.get_lead_field(l, "utm_medium"),
            "utm_campaign": ax.get_lead_field(l, "utm_campaign"),
            "Теги": ax.get_tags(l),
        })

    # запись xlsx
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сделки"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])

    hfill = PatternFill("solid", fgColor="2F5496")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")
    widths = [10, 46, 30, 22, 22, 22, 14, 16, 20, 16, 16, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    base = f"leads_dump_{date_from_dt:%Y-%m-%d}_{date_to_dt:%Y-%m-%d}"
    fpath = os.path.join(OUT_DIR, base + ".xlsx")
    wb.save(fpath)

    # CSV-двойник (utf-8-sig, чтобы корректно открывался в Excel и читался программно)
    import csv
    cpath = os.path.join(OUT_DIR, base + ".csv")
    with open(cpath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for r in rows:
            w.writerow([r.get(c, "") for c in COLUMNS])

    # быстрая прикидка по заполненности
    total = len(rows)
    no_src = sum(1 for r in rows if not str(r["Источник"]).strip())
    no_tags = sum(1 for r in rows if not str(r["Теги"]).strip())
    print(f"\nЗаписано: {fpath}")
    print(f"          {cpath}")
    print(f"Всего сделок: {total}")
    if total:
        print(f"Без источника: {no_src} ({no_src/total*100:.1f}%)")
        print(f"Без тегов: {no_tags} ({no_tags/total*100:.1f}%)")
    print("\nОткрой файл и посмотри данные — дальше решим, как анализировать.")


if __name__ == "__main__":
    main()
